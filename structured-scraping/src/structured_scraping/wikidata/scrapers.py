"""High-level scraper functions for common PEP data extraction tasks.

This module provides simple, high-level interfaces for common PEP scraping
tasks without requiring users to understand SPARQL or low-level details.
"""

import datetime as dt
import hashlib
import logging
import re
from collections.abc import Sequence
from concurrent.futures import ThreadPoolExecutor, as_completed
from numbers import Real
from pathlib import Path
from time import perf_counter, sleep
from urllib.parse import unquote, urlparse

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from pandas import ExcelWriter

from structured_scraping.sparql_utils import (
    count_sparql_query,
    extract_bindings,
    query_sparql_endpoint_with_retry,
    resilient_batched_sparql_query,
    save_results_to_csv,
)

from . import DEFAULT_USER_AGENT, WDQS_ENDPOINT
from .countries import get_country_id, get_country_name
from .filters import filter_relevant_peps
from .queries.pep import (
    ALIAS_POLITICIANS_QUERY,
    CRIMINAL_POLITICIANS_QUERY,
    DOB_POLITICIANS_QUERY,
    MAIN_QUERY,
    NATIONALITY_POLITICIANS_QUERY,
    RCA_POLITICIANS_QUERY,
    RESIDENCE_POLITICIANS_QUERY,
    ROLE_POLITICIANS_QUERY,
)

logger = logging.getLogger(__name__)


class QueryProgressTracker:
    """Track and log progress for individual SPARQL query attempts.

    Attributes:
        description (str): Context label for log messages.
        total_attempts (int): Estimated number of attempts expected.
        completed_attempts (int): Count of attempts already executed.
        durations (list[float]): Recorded durations for completed attempts.

    """

    description: str
    total_attempts: int
    completed_attempts: int
    durations: list[float]

    def __init__(self, total_attempts: int, description: str) -> None:
        """Initialise the tracker with an optional attempt estimate.

        Args:
            total_attempts (int): Anticipated number of SPARQL attempts.
            description (str): Context label for log messages.

        """
        self.description = description
        self.total_attempts = max(total_attempts, 0)
        self.completed_attempts = 0
        self.durations = []

    def record_attempt(self, duration: float) -> None:
        """Record the duration of a completed SPARQL attempt and log progress.

        Args:
            duration (float): Elapsed time for the completed attempt in seconds.

        Returns:
            None

        """
        self.completed_attempts += 1
        self.durations.append(duration)
        self.total_attempts = max(self.total_attempts, self.completed_attempts)

        _log_query_progress(
            self.completed_attempts,
            self.total_attempts,
            self.durations,
            self.description,
        )


def _format_query_progress_bar(
    completed: int,
    total: int,
    width: int = 30,
) -> str:
    """Create an ASCII progress bar showing completed query count.

    Args:
        completed (int): Number of queries that have finished running.
        total (int): Total number of queries scheduled.
        width (int, optional): Number of character cells used for the bar.
            Defaults to 30.

    Returns:
        str: Formatted progress bar string such as ``[=====-----]``.

    """
    if total <= 0 or width <= 0:
        return "[" + "-" * max(width, 0) + "]"

    fraction_complete = max(0.0, min(1.0, completed / total))
    filled_cells = round(fraction_complete * width)
    filled_cells = min(width, max(0, filled_cells))
    empty_cells = width - filled_cells
    return f"[{'=' * filled_cells}{'-' * empty_cells}]"


def _log_query_progress(
    completed: int,
    total: int,
    durations: Sequence[float],
    description: str = "Query progress",
) -> None:
    """Log a progress bar with an estimated duration for upcoming queries.

    Args:
        completed (int): Count of queries that have completed.
        total (int): Total number of queries scheduled.
        durations (Sequence[float]): Durations (in seconds) for completed queries.
        description (str, optional): Context label for the progress message.
            Defaults to ``"Query progress"``.

    Returns:
        None

    """
    progress_bar = _format_query_progress_bar(completed, total)
    remaining = max(0, total - completed)
    if not durations:
        eta_seconds = 0.0
    else:
        average_duration = sum(durations) / len(durations)
        eta_seconds = average_duration

    if remaining == 0:
        logger.info(
            "%s %s %d/%d complete. All queries processed.",
            description,
            progress_bar,
            completed,
            total,
        )
        return

    logger.info(
        "%s %s %d/%d complete. Estimated time for next query: %.2f seconds",
        description,
        progress_bar,
        completed,
        total,
        eta_seconds,
    )


def _estimate_decade_attempts(start_decade: int, end_decade: int) -> int:
    """Estimate the number of SPARQL attempts for decade-based querying.

    Args:
        start_decade (int): The starting decade of the range.
        end_decade (int): The final decade of the range.

    Returns:
        int: Estimated number of SPARQL attempts required when falling back to
        smaller windows in the worst case.

    """
    if end_decade < start_decade:
        return 1

    decades = ((end_decade - start_decade) // 10) + 1
    worst_case_per_decade = 21  # Full decade + half-decade + two-year + one-year fallbacks
    return 1 + decades * worst_case_per_decade


class PEPScraperConfig:
    """Configuration for PEP scraping operations.
    
    Attributes:
        batch_size (int): Number of results to fetch per batch.
        pause_s (float): Pause between batches in seconds.
        max_retries (int): Maximum retry attempts for rate limiting.
        timeout (int): Query timeout in seconds.
        language (str): Language code for labels.
        use_batching (bool): Whether to use batched queries for large datasets.
        use_decades (bool): Whether do batch queries by decade of birth
        
    """
    
    def __init__(
        self,
        batch_size: int = 3000,
        pause_s: float = 2.0,
        max_retries: int = 5,
        timeout: int = 60,
        language: str = "en",
        use_batching: bool = True,
        use_decades: bool = True,
    ) -> None:
        """Initialize PEP scraper configuration.
        
        Args:
            batch_size (int, optional): Number of results per batch. Defaults to 3000.
            pause_s (float, optional): Pause between batches. Defaults to 2.0.
            max_retries (int, optional): Maximum retries. Defaults to 5.
            timeout (int, optional): Query timeout. Defaults to 60.
            language (str, optional): Language for labels. Defaults to "en".
            use_batching (bool, optional): Use batched queries. Defaults to True.
            use_decades (bool, optional): Use decade-based filtering. Defaults to True.

        """
        self.batch_size = batch_size
        self.pause_s = pause_s
        self.max_retries = max_retries
        self.timeout = timeout
        self.language = language
        self.use_batching = use_batching
        self.use_decades = use_decades


def count_country_politicians(
    country: str,
    living_only: bool = False,
    config: PEPScraperConfig | None = None,
) -> int:
    """Count politicians from a specific country.
    
    Args:
        country (str): Country code (e.g., 'uk', 'us') or Wikidata ID.
        living_only (bool, optional): Only count living politicians. Defaults to False.
        config (PEPScraperConfig | None, optional): Configuration. Defaults to None.
        
    Returns:
        int: Total number of politicians found.
        
    Raises:
        ValueError: If country is not recognized.
        Exception: If the count query fails.
        
    """
    if config is None:
        config = PEPScraperConfig()
    
    country_id = get_country_id(country)
    country_name = get_country_name(country)
    
    logger.info("Counting politicians from %s (%s)", country_name, country_id)
    
    # Build COUNT query based on whether we want living politicians only
    if living_only:
        # COUNT query for living politicians only
        count_query = f"""
SELECT (COUNT(DISTINCT ?person) AS ?count) WHERE {{
  ?person wdt:P31 wd:Q5 .        # Instance of human
  ?person wdt:P106 wd:Q82955 .   # Occupation: politician
  ?person wdt:P27 wd:{country_id} .  # Country of citizenship
  
  # Filter out deceased politicians (no death date)
  FILTER NOT EXISTS {{ ?person wdt:P570 ?deathDate . }}
}}
"""
    else:
        # COUNT query for all politicians (living and deceased)
        count_query = f"""
SELECT (COUNT(DISTINCT ?person) AS ?count) WHERE {{
  ?person wdt:P31 wd:Q5 .        # Instance of human
  ?person wdt:P106 wd:Q82955 .   # Occupation: politician
  ?person wdt:P27 wd:{country_id} .  # Country of citizenship
}}
"""
    
    return count_sparql_query(
        endpoint_url=WDQS_ENDPOINT,
        query=count_query,
        timeout=config.timeout,
        user_agent=DEFAULT_USER_AGENT,
    )

def scrape_country_politicians(
    country: str,
    output_file: str | Path | None = None,
    living_only: bool = False,
    apply_relevance_filter: bool = True,
    config: PEPScraperConfig | None = None,
) -> tuple[int, str]:
    """Scrape politicians from a specific country and save to CSV.
    
    Args:
        country (str): Country code (e.g., 'uk', 'us') or Wikidata ID.
        output_file (str | Path | None, optional): Output CSV file path.
            If None, auto-generates timestamped filename.
        living_only (bool, optional): Only scrape living politicians. Defaults to False.
        apply_relevance_filter (bool, optional): Apply post-query filtering. Defaults to True.
        config (PEPScraperConfig | None, optional): Configuration. Defaults to None.
        
    Returns:
        tuple[int, str]: Number of records scraped and output file path.
        
    Raises:
        ValueError: If country is not recognized.
        Exception: If scraping fails.
        
    """
    if config is None:
        config = PEPScraperConfig()
    
    country_id = get_country_id(country)
    country_name = get_country_name(country)
    
    logger.info(
        "Scraping %s politicians from %s (%s)",
        "living" if living_only else "all",
        country_name,
        country_id,
    )
    
    # Generate output filename if not provided
    if output_file is None:
        timestamp = dt.datetime.now(tz=dt.UTC).strftime("%Y%m%d_%H%M%S")
        living_suffix = "_living" if living_only else ""
        relevant_suffix = "_relevant" if apply_relevance_filter else ""
        output_file = f"pep_{country_name.lower()}{living_suffix}{relevant_suffix}_{timestamp}.csv"
    
    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Use the nationality-based query template
    base_query = MAIN_QUERY.strip()
    
    # Choose appropriate query based on living_only flag
    if living_only:
        # Use specialized living query to avoid SPARQL optimization issues
        query = MAIN_QUERY.replace("nationality_qid",country_id)
    else:
        # Use the standard query for all politicians (living and deceased)
        query = base_query.replace(
            "nationality_qid",country_id,
        )
    
    logger.info("Executing query...")
    

    
    if config.use_batching:
        # Use batched execution for large datasets
        logger.info("Using batched query execution")
        all_results, skipped_records = resilient_batched_sparql_query(
            endpoint_url=WDQS_ENDPOINT,
            base_query=query,
            batch_size=config.batch_size,
            pause_s=config.pause_s,
            timeout=config.timeout,
            max_retries=config.max_retries,
            user_agent=DEFAULT_USER_AGENT,
        )
        
        if skipped_records > 0:
            logger.warning("Skipped %d corrupted records", skipped_records)
    else:
        # Use simple query execution (no batching)
        logger.info("Using simple query execution (no batching)")
        sparql_results = query_sparql_endpoint_with_retry(
            endpoint_url=WDQS_ENDPOINT,
            query=query,
            return_format="json",
            timeout=config.timeout,
            user_agent=DEFAULT_USER_AGENT,
        )
        
        all_results = extract_bindings(sparql_results)
        skipped_records = 0
    
    # Apply relevance filter if requested and not already living-only
    if apply_relevance_filter and not living_only:
        logger.info("Applying relevance filter...")
        filtered_results = filter_relevant_peps(all_results)
        logger.info(
            "Filtered from %d to %d relevant records",
            len(all_results),
            len(filtered_results),
        )
        final_results = filtered_results
    else:
        final_results = all_results
    
    # Sort results by person name
    logger.info("Sorting results...")
    sorted_results = sorted(
        final_results,
        key=lambda x: x.get("personLabel", "").lower(),
    )
    
    # Save to CSV - let save_results_to_csv automatically detect all columns
    # This ensures all scraped data fields are included in the output
    logger.info("Saving %d results to %s", len(sorted_results), output_path)
    save_results_to_csv(
        results=sorted_results,
        file_path=output_path,
        fieldnames=None,  # Auto-detect all columns from the data
    )
    
    return len(sorted_results), str(output_path)



def scrape_country_politicians_by_decade(  # noqa: C901, PLR0912, PLR0915
    country: str,
    output_file: str | Path | None = None,
    living_only: bool = False,
    apply_relevance_filter: bool = True,
    config: PEPScraperConfig | None = None,
    start_decade: int = 1920,
    end_decade: int = 2020,
) -> tuple[int, str]:
    """Scrape politicians from a specific country and save to CSV.
    
    Args:
        country (str): Country code (e.g., 'uk', 'us') or Wikidata ID.
        output_file (str | Path | None, optional): Output CSV file path.
            If None, auto-generates timestamped filename.
        living_only (bool, optional): Only scrape living politicians. Defaults to False.
        apply_relevance_filter (bool, optional): Apply post-query filtering. Defaults to True.
        config (PEPScraperConfig | None, optional): Configuration. Defaults to None.
        start_decade (int): The starting decade of the birth date filter.
        end_decade (int): The ending decade of the birth date filter.

    Returns:
        tuple[int, str]: Number of records scraped and output file path.
        
    Raises:
        ValueError: If country is not recognized.
        Exception: If scraping fails.
        
    """
    logger.info("USING CORRECT DEFINITION")
    run_started_at = dt.datetime.now(tz=dt.UTC)
    start_time = perf_counter()
    if config is None:
        config = PEPScraperConfig()
    
    
    country_id = get_country_id(country)
    country_name = get_country_name(country)
    
    logger.info(
        "Scraping %s politicians from %s (%s)",
        "living" if living_only else "all",
        country_name,
        country_id,
    )
    
    # Generate output filename if not provided
    if output_file is None:
        timestamp = run_started_at.strftime("%Y%m%d_%H%M%S")
        living_suffix = "_living" if living_only else ""
        relevant_suffix = "_relevant" if apply_relevance_filter else ""
        output_file = f"pep_{country_name.lower()}{living_suffix}{relevant_suffix}_{timestamp}.xlsx"

    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    query_list = [
        ("Main", MAIN_QUERY),
        ("DOB", DOB_POLITICIANS_QUERY),
        ("Nationality", NATIONALITY_POLITICIANS_QUERY),
        ("Alias", ALIAS_POLITICIANS_QUERY),
        ("Address", RESIDENCE_POLITICIANS_QUERY),
        ("Case Details", CRIMINAL_POLITICIANS_QUERY),
        ("Role Type", ROLE_POLITICIANS_QUERY),
        ("RCA", RCA_POLITICIANS_QUERY),
    ]
    
    results_by_query: dict[str, list[dict[str, str]]] = {}
    collected_ids = []
    total_queries = len(query_list)
    query_durations: list[float] = []
    
    for query_index, (query_name, query) in enumerate(query_list, start=1):
        all_results: list[dict[str, str]] = []

        logger.info("Running query: %s", query_name)
        query_start_time = perf_counter()
        tracker_label = f"{query_name} SPARQL attempts"
        if config.use_decades:
            decade_attempts = _estimate_decade_attempts(start_decade, end_decade)
            if query == MAIN_QUERY:
                tracker = QueryProgressTracker(decade_attempts, tracker_label)
            else:
                tracker = QueryProgressTracker(max(len(dict.fromkeys(collected_ids)), 1), tracker_label)

            logger.info("IF LOOP WORKED")

            if query == MAIN_QUERY:
                results = decade_loop(
                    country,
                    query,
                    start_decade,
                    end_decade,
                    config=config,
                    tracker=tracker,
                )
                if results:
                    all_results.extend(results)
                    ids = [extract_qid(r) for r in results if extract_qid(r)]
                    collected_ids.extend(ids)
            else:
                results = attempt_period(
                    country,
                    query,
                    collected_ids=collected_ids,
                    config=config,
                    tracker=tracker,
                    start_year=0,
                    end_year=0,
                )

                if results:
                    all_results.extend(results)

        elif query == MAIN_QUERY:
            tracker = QueryProgressTracker(2, tracker_label)
            results = attempt_period(country, query, 0, 0, config=config, tracker=tracker)
            candidate_results: list[dict[str, str]] = []
            if results is not None:
                candidate_results.extend(results)
                ids = [extract_qid(r) for r in results if extract_qid(r)]
                collected_ids.extend(ids)
            results = attempt_period(
                country,
                query,
                start_decade,
                end_decade,
                config=config,
                tracker=tracker,
            )
            if results is not None:
                candidate_results.extend(results)
                ids = [extract_qid(r) for r in results if extract_qid(r)]
                collected_ids.extend(ids)

            if collected_ids:
                all_results.extend(candidate_results)
        else:
            tracker = QueryProgressTracker(max(len(dict.fromkeys(collected_ids)), 1), tracker_label)
            results = attempt_period(
                country,
                query,
                0,
                0,
                config=config,
                collected_ids=collected_ids,
                tracker=tracker,
            )
            if results is not None:
                all_results.extend(results)

        # Apply relevance filter if requested and not already living-only
        if apply_relevance_filter and not living_only:
            logger.info("Applying relevance filter to %s", query_name)
            filtered_results = filter_relevant_peps(all_results)
            logger.info(
                "Filtered from %d to %d relevant records",
                len(all_results),
                len(filtered_results),
            )
            final_results = filtered_results
        else:
            final_results = all_results

        if query == MAIN_QUERY and final_results:
            final_results = _enrich_main_results(final_results, config)

        # Sort results by person name
        logger.info("Sorting results...")
        sorted_results = sorted(
            final_results,
            key=lambda x: x.get("personLabel", "").lower(),
        )
        results_by_query[query_name] = sorted_results
        logger.info("Query %s collected %d records", query_name, len(sorted_results))

        query_duration = perf_counter() - query_start_time
        query_durations.append(query_duration)
        logger.info("Completed query %s in %.2f seconds", query_name, query_duration)
        _log_query_progress(query_index, total_queries, query_durations)

    if not any(results_by_query.values()):
        raise ValueError(
            "No records were collected from Wikidata, so no Excel sheets were written. "
            "Check the earlier SPARQL errors for the root cause.",
        )

    with ExcelWriter(output_path, engine="openpyxl") as writer:
        for query_name, results in results_by_query.items():
            if not results:
                continue
            df = pd.DataFrame(results)

            # 🔁 Replace specific values in 'convictedOfLabel'
            if "convictedOfLabel" in df.columns:
                df["convictedOfLabel"] = df["convictedOfLabel"].replace({
                    "Q15279749": "Range of legal and illegal activities that reduce tax paid",
                    # Add more replacements here if needed
                })

            if query_name == "Alias":
                df = _normalise_alias_columns(df)

            if not df.empty and "ID" in df.columns:
                if query_name == "Role Type":
                    df = (
                        df.groupby("ID")
                        .apply(aggregate_role_type_group)
                        .reset_index(drop=True)
                    )
                elif query_name == "Main":
                    df = df.groupby("ID", as_index=False).agg(_aggregate_main_column)
                elif query_name == "Alias":
                    df = df.groupby("ID", as_index=False).agg(_aggregate_alias_column)
                elif query_name == "RCA":
                    df = df.groupby("ID", as_index=False).agg(
                        lambda x: x.iloc[0] if x.name == "ID" else wrap_values(x),
                    )
                else:
                    df = df.groupby("ID", as_index=False).agg(
                        lambda x: wrap_values(x) if x.name != "ID" else x.iloc[0],
                    )
            df = bring_id_to_front(df)
            sheet_name = query_name[:31]
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            logger.info("Wrote %d records to sheet: %s", len(df), sheet_name)

    apply_main_sheet_hyperlinks(output_path)

    total_records = sum(len(v) for v in results_by_query.values())
    run_completed_at = dt.datetime.now(tz=dt.UTC)
    duration = perf_counter() - start_time
    _append_run_metadata_sheet(
        output_path=output_path,
        country_name=country_name,
        run_started_at=run_started_at,
        run_completed_at=run_completed_at,
        duration_seconds=duration,
        total_records=total_records,
    )

    logger.info(
        "scrape_country_politicians_by_decade finished in %.2f seconds",
        duration,
    )
    return total_records, str(output_path)

def _resolve_image_redirect(url: str) -> str:
    """Convert a Wikimedia redirect-style image URL into a direct file URL.

    Args:
        url (str): URL value retrieved from Wikidata.

    Returns:
        str: Direct image URL if conversion succeeds, otherwise the original URL.

    """
    try:
        parsed_url = urlparse(url)
    except ValueError:
        return url
    marker = "Special:FilePath/"
    if not parsed_url.path or marker not in parsed_url.path:
        return url
    filename = parsed_url.path.split(marker, maxsplit=1)[1]
    if not filename:
        return url
    filename = unquote(filename)
    if not filename:
        return url
    filename = filename.split("/", maxsplit=1)[0]
    filename = filename.replace(" ", "_")
    filename = filename.removeprefix("File:")
    if not filename:
        return url
    digest = hashlib.md5(filename.encode("utf-8")).hexdigest()  # noqa: S324 # Required by Wikimedia file path scheme
    return f"https://upload.wikimedia.org/wikipedia/commons/{digest[0]}/{digest[:2]}/{filename}"


def select_first_non_empty_value(series: pd.Series, column_name: str | None = None) -> str:
    """Return the first non-empty value from a Series as a string.

    Args:
        series (pd.Series): Series containing candidate values.
        column_name (str | None, optional): Column name used for special handling.

    Returns:
        str: First non-empty value if present, otherwise an empty string.

    """
    for value in series:
        if isinstance(value, str):
            candidate = value.strip()
            if not candidate:
                continue
            if column_name == "image":
                return _resolve_image_redirect(candidate)
            return candidate
        if pd.notna(value):
            candidate = str(value).strip()
            if not candidate:
                continue
            if column_name == "image":
                return _resolve_image_redirect(candidate)
            return candidate
    return ""

def wrap_values(x: pd.Series, column_name: str | None = None) -> str:
    """Convert a pandas Series to a comma-separated string of unique values, each wrapped in brackets.

    Args:
        x (pd.Series): A pandas Series of values to process.
        column_name (str | None, optional): Name of the column being processed.

    Returns:
        str: A formatted string of unique, bracketed values.

    """
    if column_name == "endTime":
        values = x.astype(str).fillna("None")
        unique_vals = sorted({val if val.lower() != "nan" else "None" for val in values})
    else:
        values = x.astype(str)
        unique_vals = sorted({val for val in values if val.lower() != "nan" and val != ""})
    # Only wrap non-empty values
    return ", ".join(f'["{val}"]' for val in unique_vals if val)

def strip_outer_brackets(aka: str) -> str:
    """Remove leading '["' and trailing '"]' from each name in a bracketed, comma-separated string.

    Args:
        aka (str): The cell value containing bracketed names.

    Returns:
        str: Comma-separated names without outer brackets.

    """
    if not aka or aka.lower() == "none":
        return ""
    # Extract names inside brackets
    names = re.findall(r'\["(.*?)"\]', aka)
    if not names:
        # If not bracketed, split by comma and strip whitespace/brackets/quotes
        names = [n.strip(' []"') for n in aka.split(",") if n.strip()]
    return ", ".join(names)


def _split_alias_values(value: str) -> list[str]:
    """Split the contents of an Alias cell into individual names.

    Args:
        value (str): Raw cell value from the Alias sheet.

    Returns:
        list[str]: Cleaned list of individual names extracted from the value.

    """
    if not isinstance(value, str):
        return []
    if not value or value.lower() == "none":
        return []
    bracketed = re.findall(r'\["(.*?)"\]', value)
    if bracketed:
        return [name.strip() for name in bracketed if name.strip()]
    parts = re.split(r"[;,]", value)
    names = [part.strip(' []"') for part in parts if part.strip(' []"')]
    if names:
        return names
    stripped = value.strip()
    return [stripped] if stripped else []


def _aggregate_main_column(series: pd.Series) -> object:
    """Aggregate Main sheet columns with custom handling for IDs and images.

    Args:
        series (pd.Series): Grouped column values for a specific Main sheet record.

    Returns:
        object: Aggregated value ready for export.

    """
    column_name = str(series.name) if series.name is not None else None
    if column_name == "ID":
        return series.iloc[0]
    if column_name in {"person", "image"}:
        return select_first_non_empty_value(series, column_name=column_name)
    return wrap_values(series, column_name=column_name)


def _aggregate_alias_column(series: pd.Series) -> object:
    """Aggregate Alias sheet columns into bracketed, unique values.

    Args:
        series (pd.Series): Grouped column values for a specific Alias record.

    Returns:
        object: Aggregated value ready for export.

    """
    column_name = str(series.name) if series.name is not None else None
    if column_name == "ID":
        return series.iloc[0]
    return wrap_values(series, column_name=column_name)


def _normalise_alias_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Clean Alias sheet columns and drop unsupported original-script data.

    Args:
        df (pd.DataFrame): Alias sheet DataFrame.

    Returns:
        pd.DataFrame: Normalised DataFrame ready for export.

    """
    alias_df = df.copy()
    if "aka" in alias_df.columns:
        alias_df["aka"] = alias_df["aka"].apply(strip_outer_brackets)
    return (
        alias_df.drop(columns=["Original Script Name"])
        if "Original Script Name" in alias_df.columns
        else alias_df
    )


def _normalise_cell_value(
    value: str | bytes | Real | None,
) -> str | None:
    """Return a cleaned string for a cell value or None when empty.

    Args:
        value (str | bytes | float | int | None): Raw cell value.

    Returns:
        str | None: Stripped string value, or None if blank/NaN.

    """
    # type: ignore is used here because pd.isna accepts several scalar types, but not generic object
    if pd.isna(value):  # type: ignore # Acceptable types for pd.isna are provided in the annotation
        return None
    value_str = str(value).strip()
    if not value_str or value_str.lower() == "nan":
        return None
    return value_str


def _collect_unique_starts(label_rows: pd.DataFrame) -> list[str]:
    """Extract unique, ordered start times from a role-specific subset.

    Args:
        label_rows (pd.DataFrame): Rows for a single position label, including start times.

    Returns:
        list[str]: Unique start times in the order they first appear.

    """
    starts: list[str] = []
    seen_starts: set[str] = set()
    for raw_start in label_rows["startTime"]:
        normalised_start = _normalise_cell_value(raw_start)
        if normalised_start is None or normalised_start in seen_starts:
            continue
        seen_starts.add(normalised_start)
        starts.append(normalised_start)
    return starts


def _format_end_entries(label_rows: pd.DataFrame, starts: list[str]) -> list[str]:
    """Create formatted end-time entries aligned to provided starts.

    Args:
        label_rows (pd.DataFrame): Rows for a single position label containing potential end times.
        starts (list[str]): Normalised start times requiring matching end values.

    Returns:
        list[str]: Bracketed end-time strings aligned with ``starts``.

    """
    formatted: list[str] = []
    has_end_time = "endTime" in label_rows.columns
    for start in starts:
        end_value = "None"
        if has_end_time:
            matching_end = label_rows.loc[label_rows["startTime"] == start, "endTime"]
            for candidate in matching_end:
                candidate_norm = _normalise_cell_value(candidate)
                if candidate_norm is not None:
                    end_value = candidate_norm
                    break
        formatted.append(f'"{end_value}"')
    return [f"[{entry}]" for entry in formatted]


def build_role_type_timeline(group: pd.DataFrame) -> tuple[list[str], list[str], list[str]]:
    """Construct aligned label, start, and end lists for a role group.

    Args:
        group (pd.DataFrame): Role Type rows belonging to one person.

    Returns:
        tuple[list[str], list[str], list[str]]: Bracketed position labels, start times, and end times.

    """
    label_entries: list[str] = []
    start_entries: list[str] = []
    end_entries: list[str] = []
    if "positionLabel" not in group.columns or "startTime" not in group.columns:
        return label_entries, start_entries, end_entries
    for label_value, label_rows in group.groupby("positionLabel"):
        label_str = _normalise_cell_value(str(label_value))
        if label_str is None:
            continue
        ordered_starts = _collect_unique_starts(label_rows)
        if not ordered_starts:
            continue
        label_entries.extend([f'["{label_str}"]' for _ in ordered_starts])
        start_entries.extend([f'["{start}"]' for start in ordered_starts])
        end_entries.extend(_format_end_entries(label_rows, ordered_starts))
    return label_entries, start_entries, end_entries


def aggregate_role_type_group(group: pd.DataFrame) -> pd.Series:
    """Aggregate Role Type rows so designations align with start dates.

    Args:
        group (pd.DataFrame): Role Type rows belonging to a single person.

    Returns:
        pd.Series: Aggregated values for the Role Type sheet.

    """
    aggregated: dict[str, str] = {"ID": group["ID"].iloc[0]}
    label_entries, start_entries, end_entries = build_role_type_timeline(group)
    if label_entries:
        aggregated["positionLabel"] = ", ".join(label_entries)
        aggregated["startTime"] = ", ".join(start_entries)
        aggregated["endTime"] = ", ".join(end_entries)
    else:
        if "positionLabel" in group.columns:
            aggregated["positionLabel"] = wrap_values(group["positionLabel"], column_name="positionLabel")
        if "startTime" in group.columns:
            aggregated["startTime"] = wrap_values(group["startTime"], column_name="startTime")
        if "endTime" in group.columns:
            aggregated["endTime"] = wrap_values(group["endTime"], column_name="endTime")
    for column in group.columns:
        if column in {"ID", "positionLabel", "startTime", "endTime"}:
            continue
        aggregated[column] = wrap_values(group[column], column_name=column)
    return pd.Series(aggregated)


def decade_loop(
    country: str,
    query: str,
    start_decade: int,
    end_decade: int,
    collected_ids: list[str] | None = None,
    config: PEPScraperConfig | None = None,
    tracker: QueryProgressTracker | None = None,
    include_missing_birth_date: bool = True,
) -> list[dict[str, str]]:
    """Execute a query across decade windows with progressively smaller retries.

    Args:
        country (str): Country code (e.g., 'uk', 'us') or Wikidata ID.
        query (str): Template of the SPARQL query to execute.
        start_decade (int): The starting decade of the birth date filter.
        end_decade (int): The ending decade of the birth date filter.
        collected_ids (list[str] | None): Optional list of QIDs scoped to the query.
        config (PEPScraperConfig | None): Optional scraper configuration overrides.
        tracker (QueryProgressTracker | None): Optional progress tracker for attempts.
        include_missing_birth_date (bool): Whether to run the expensive
            no-birth-date window before decade windows.

    Returns:
        list[dict[str, str]]: Aggregated records across all attempted windows.

    Raises:
        None.

    """
    all_results: list[dict[str, str]] = []
    active_config = config or PEPScraperConfig()

    # Search for those without a date of birth only when explicitly requested.
    if include_missing_birth_date:
        results = attempt_period(
            country,
            query,
            0,
            0,
            config=active_config,
            collected_ids=collected_ids,
            tracker=tracker,
        )
        if results is not None:
            all_results.extend(results)
        
    # Loop over each decade, with retries using smaller ranges if needed
    for decade in range(start_decade, end_decade + 1, 10):
        logger.info("Querying for people born between %d and %d", decade, decade + 10)

        results = attempt_period(
            country,
            query,
            decade,
            decade + 10,
            config=active_config,
            collected_ids=collected_ids,
            tracker=tracker,
        )

        if results is not None:
            all_results.extend(results)
            continue

        # Retry with 5-year intervals
        for half_decade in [decade, decade + 5]:
            results = attempt_period(
                country,
                query,
                half_decade,
                half_decade + 5,
                config=active_config,
                collected_ids=collected_ids,
                tracker=tracker,
            )
            if results is not None:
                all_results.extend(results)
                continue

            # Retry with 2-year intervals
            for two_year in range(half_decade, half_decade + 5, 2):
                results = attempt_period(
                    country,
                    query,
                    two_year,
                    min(two_year + 2, half_decade + 5),
                    config=active_config,
                    collected_ids=collected_ids,
                    tracker=tracker,
                )
                if results is not None:
                    all_results.extend(results)
                    continue

                # Retry with 1-year intervals
                for year in range(two_year, min(two_year + 2, half_decade + 5)):
                    results = attempt_period(
                        country,
                        query,
                        year,
                        year + 1,
                        config=active_config,
                        collected_ids=collected_ids,
                        tracker=tracker,
                    )
                    if results is not None:
                        all_results.extend(results)
    return all_results

def apply_main_sheet_hyperlinks(output_path: Path) -> None:
    """Convert plain URL strings in the Main sheet into clickable hyperlinks.

    Args:
        output_path (Path): Path to the generated Excel workbook.

    """
    workbook = load_workbook(output_path)
    if "Main" not in workbook.sheetnames:
        workbook.close()
        return

    worksheet = workbook["Main"]
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    person_idx = headers.index("person") + 1 if "person" in headers else None
    image_idx = headers.index("image") + 1 if "image" in headers else None

    for row_idx in range(2, worksheet.max_row + 1):
        for col_idx in (person_idx, image_idx):
            if col_idx is None:
                continue
            cell = worksheet.cell(row=row_idx, column=col_idx)
            url = cell.value
            # Only assign hyperlink if cell is not a MergedCell
            if isinstance(cell, Cell) and isinstance(url, str) and url.startswith("http"):
                cell.value = url
                cell.hyperlink = url
                cell.style = "Hyperlink"

    workbook.save(output_path)
    workbook.close()


def _append_run_metadata_sheet(
    output_path: Path,
    country_name: str,
    run_started_at: dt.datetime,
    run_completed_at: dt.datetime,
    duration_seconds: float,
    total_records: int,
) -> None:
    """Append or refresh a summary sheet containing scraper runtime metadata.

    Args:
        output_path (Path): Path to the generated Excel workbook.
        country_name (str): Human-readable country name for context.
        run_started_at (datetime): UTC timestamp when the scrape began.
        run_completed_at (datetime): UTC timestamp when the scrape ended.
        duration_seconds (float): Total runtime in seconds.
        total_records (int): Aggregate number of exported records.

    Returns:
        None

    """
    workbook = load_workbook(output_path)
    sheet_name = "Run Info"
    if sheet_name in workbook.sheetnames:
        existing_sheet = workbook[sheet_name]
        workbook.remove(existing_sheet)
    info_sheet = workbook.create_sheet(title=sheet_name)

    info_sheet.append(["Metric", "Value"])
    summary_rows = [
        ("Country", country_name),
        ("Run started at (UTC)", run_started_at.isoformat()),
        ("Run completed at (UTC)", run_completed_at.isoformat()),
    ("Elapsed (seconds)", f"{duration_seconds:.2f}"),
    ("Elapsed (hh:mm:ss)", str(dt.timedelta(seconds=duration_seconds))),
        ("Total records exported", str(total_records)),
    ]
    for metric, value in summary_rows:
        info_sheet.append([metric, value])

    workbook.save(output_path)
    workbook.close()

def extract_qid(row: dict) -> str | None:
    """Extract the Q-code from a result row safely."""
    if "ID" in row:
        qid = row["ID"]
        if qid.startswith("Q") and qid[1:].isdigit():
            return qid
    if "person" in row:
        return row["person"].split("/")[-1]
    return None


def _execute_main_query(
    country: str,
    base_query: str,
    start_year: int,
    end_year: int,
    config: PEPScraperConfig,
    tracker: QueryProgressTracker | None,
) -> list[dict[str, str]] | None:
    """Execute the main query across a specific date window.

    Args:
        country (str): The country code or QID to filter results by nationality.
        base_query (str): The SPARQL query template to execute.
        start_year (int): The starting year of the birth date filter.
        end_year (int): The ending year of the birth date filter.
        config (PEPScraperConfig): Active scraper configuration values.
        tracker (QueryProgressTracker | None): Optional progress tracker for attempts.

    Returns:
        list[dict[str, str]] | None: Extracted rows when the query succeeds, else ``None``.

    Raises:
        None.

    """
    results: list[dict[str, str]] = []
    query = main_query_format(start_year, end_year, base_query, country)
    attempt_start = perf_counter()
    try:
        if config.use_batching:
            logger.info("Using batched query execution")
            results, skipped_records = resilient_batched_sparql_query(
                endpoint_url=WDQS_ENDPOINT,
                base_query=query,
                batch_size=config.batch_size,
                pause_s=config.pause_s,
                timeout=config.timeout,
                max_retries=config.max_retries,
                user_agent=DEFAULT_USER_AGENT,
            )
            if skipped_records > 0:
                logger.warning("Skipped %d corrupted records", skipped_records)
        else:
            logger.info("Using simple query execution")
            sparql_results = query_sparql_endpoint_with_retry(
                endpoint_url=WDQS_ENDPOINT,
                query=query,
                return_format="json",
                timeout=config.timeout,
                user_agent=DEFAULT_USER_AGENT,
            )
            results = extract_bindings(sparql_results)
    except Exception as exc:  # noqa: BLE001 # The SPARQL client raises generic Exception on HTTP or parsing failures
        duration = perf_counter() - attempt_start
        if tracker is not None:
            tracker.record_attempt(duration)
        logger.warning("Query failed for %d-%d: %s", start_year, end_year, str(exc))
        return None

    duration = perf_counter() - attempt_start
    if tracker is not None:
        tracker.record_attempt(duration)
    return results


def _execute_person_queries(
    base_query: str,
    config: PEPScraperConfig,
    collected_ids: list[str],
    tracker: QueryProgressTracker | None,
) -> list[dict[str, str]]:
    """Execute per-person queries derived from the base template.

    Args:
        base_query (str): The SPARQL query template to execute.
        config (PEPScraperConfig): Active scraper configuration values.
        collected_ids (list[str]): QIDs to substitute into the query template.
        tracker (QueryProgressTracker | None): Optional progress tracker for attempts.

    Returns:
        list[dict[str, str]]: Combined results for all executed person-specific queries.

    Raises:
        None.

    """
    results: list[dict[str, str]] = []
    person_ids = list(dict.fromkeys(collected_ids))
    if base_query == ALIAS_POLITICIANS_QUERY:
        return _execute_alias_queries(config, person_ids, tracker)

    parallel_batch_size = 5

    def run_single_person_query(person_id: str) -> tuple[str, list[dict[str, str]], float, Exception | None]:
        attempt_start = perf_counter()
        query = base_query.replace("person_qid", person_id)
        logger.debug("Running basic query for person ID: %s", person_id)
        try:
            logger.info("Using simple query execution")
            sparql_results = query_sparql_endpoint_with_retry(
                endpoint_url=WDQS_ENDPOINT,
                query=query,
                return_format="json",
                timeout=config.timeout,
                user_agent=DEFAULT_USER_AGENT,
            )
        except Exception as exc:  # noqa: BLE001 # Upstream retry helper surfaces all request errors as Exception
            duration = perf_counter() - attempt_start
            return person_id, [], duration, exc
        individual_results = extract_bindings(sparql_results)
        duration = perf_counter() - attempt_start
        return person_id, individual_results, duration, None

    for batch_start in range(0, len(person_ids), parallel_batch_size):
        batch_ids = person_ids[batch_start : batch_start + parallel_batch_size]
        with ThreadPoolExecutor(max_workers=len(batch_ids)) as executor:
            futures = [executor.submit(run_single_person_query, person_id) for person_id in batch_ids]
            for future in as_completed(futures):
                person_id, individual_results, duration, error = future.result()
                if tracker is not None:
                    tracker.record_attempt(duration)
                if error is not None:
                    logger.warning("Query failed for person ID %s: %s", person_id, str(error))
                    continue
                if individual_results:
                    results.extend(individual_results)
        next_batch_start = batch_start + parallel_batch_size
        if config.pause_s > 0 and next_batch_start < len(person_ids):
            sleep(config.pause_s)
    return results


def _enrich_main_results(
    main_results: list[dict[str, str]],
    config: PEPScraperConfig,
) -> list[dict[str, str]]:
    """Add Main sheet detail fields using simple QLever-friendly batch queries.

    Args:
        main_results (list[dict[str, str]]): Main discovery rows.
        config (PEPScraperConfig): Active scraper configuration values.

    Returns:
        list[dict[str, str]]: Original rows plus detail rows for aggregation.

    """
    person_ids = list(dict.fromkeys(qid for row in main_results if (qid := extract_qid(row))))
    if not person_ids:
        return main_results

    enrichment_results: list[dict[str, str]] = []
    enrichment_batch_size = 50
    query_specs = [
        ("genderLabel", '?person wdt:P21 ?value . ?value rdfs:label ?genderLabel . FILTER (LANG(?genderLabel) = "en")'),
        (
            "birthPlaceLabel",
            '?person wdt:P19 ?value . ?value rdfs:label ?birthPlaceLabel . FILTER (LANG(?birthPlaceLabel) = "en")',
        ),
        ("fatherLabel", '?person wdt:P22 ?value . ?value rdfs:label ?fatherLabel . FILTER (LANG(?fatherLabel) = "en")'),
        (
            "educatedAtLabel",
            '?person wdt:P69 ?value . ?value rdfs:label ?educatedAtLabel . FILTER (LANG(?educatedAtLabel) = "en")',
        ),
        (
            "academicDegreeLabel",
            '?person wdt:P512 ?value . ?value rdfs:label ?academicDegreeLabel . FILTER (LANG(?academicDegreeLabel) = "en")',
        ),
        (
            "workLocationLabel",
            '?person wdt:P937 ?value . ?value rdfs:label ?workLocationLabel . FILTER (LANG(?workLocationLabel) = "en")',
        ),
        ("ownerOfLabel", '?person wdt:P1830 ?value . ?value rdfs:label ?ownerOfLabel . FILTER (LANG(?ownerOfLabel) = "en")'),
        ("affiliationString", "?person wdt:P6424 ?affiliationString ."),
        (
            "nonEnglishLabel",
            '?person rdfs:label ?nonEnglishLabel . FILTER (LANG(?nonEnglishLabel) != "en" && LANG(?nonEnglishLabel) != "")',
        ),
    ]

    for batch_start in range(0, len(person_ids), enrichment_batch_size):
        batch_ids = person_ids[batch_start : batch_start + enrichment_batch_size]
        values = " ".join(f"wd:{person_id}" for person_id in batch_ids)
        if not values:
            continue

        for variable_name, pattern in query_specs:
            query = f"""
SELECT DISTINCT
  (STRAFTER(STR(?person), "http://www.wikidata.org/entity/") AS ?ID)
  ?{variable_name}
WHERE {{
  VALUES ?person {{ {values} }}
  {pattern}
}}
"""
            try:
                sparql_results = query_sparql_endpoint_with_retry(
                    endpoint_url=WDQS_ENDPOINT,
                    query=query,
                    return_format="json",
                    timeout=config.timeout,
                    user_agent=DEFAULT_USER_AGENT,
                )
            except Exception as exc:  # noqa: BLE001 # Keep Main export useful if one enrichment field fails
                logger.warning("Main enrichment failed for %s: %s", variable_name, str(exc))
                continue

            enrichment_results.extend(extract_bindings(sparql_results))

        next_batch_start = batch_start + enrichment_batch_size
        if config.pause_s > 0 and next_batch_start < len(person_ids):
            sleep(config.pause_s)

    return [*main_results, *enrichment_results]


def _execute_alias_queries(
    config: PEPScraperConfig,
    person_ids: list[str],
    tracker: QueryProgressTracker | None,
) -> list[dict[str, str]]:
    """Fetch Alias sheet fields with simple property-batch queries.

    Args:
        config (PEPScraperConfig): Active scraper configuration values.
        person_ids (list[str]): QIDs to include in the alias lookup.
        tracker (QueryProgressTracker | None): Optional progress tracker.

    Returns:
        list[dict[str, str]]: Combined alias rows for the provided people.

    """
    results: list[dict[str, str]] = []
    alias_batch_size = 50
    query_specs = [
        ("AKA", "?person schema:alternateName ?AKA ."),
        ("AKA", "?person skos:altLabel ?AKA ."),
        ("nativeName", "?person wdt:P1559 ?nativeName ."),
        ("birthName", "?person wdt:P1477 ?birthName ."),
        (
            "nonEnglishLabel",
            '?person rdfs:label ?nonEnglishLabel . FILTER (LANG(?nonEnglishLabel) != "en" && LANG(?nonEnglishLabel) != "")',
        ),
    ]

    for batch_start in range(0, len(person_ids), alias_batch_size):
        batch_ids = person_ids[batch_start : batch_start + alias_batch_size]
        values = " ".join(f"wd:{person_id}" for person_id in batch_ids)
        if not values:
            continue

        for variable_name, pattern in query_specs:
            query = f"""
SELECT DISTINCT
  (STRAFTER(STR(?person), "http://www.wikidata.org/entity/") AS ?ID)
  ?{variable_name}
WHERE {{
  VALUES ?person {{ {values} }}
  {pattern}
}}
"""
            attempt_start = perf_counter()
            try:
                sparql_results = query_sparql_endpoint_with_retry(
                    endpoint_url=WDQS_ENDPOINT,
                    query=query,
                    return_format="json",
                    timeout=config.timeout,
                    user_agent=DEFAULT_USER_AGENT,
                )
            except Exception as exc:  # noqa: BLE001 # Keep alias scrape resilient if one property fails
                duration = perf_counter() - attempt_start
                if tracker is not None:
                    tracker.record_attempt(duration)
                logger.warning("Alias query failed for %s: %s", variable_name, str(exc))
                continue

            duration = perf_counter() - attempt_start
            if tracker is not None:
                tracker.record_attempt(duration)
            results.extend(extract_bindings(sparql_results))

        next_batch_start = batch_start + alias_batch_size
        if config.pause_s > 0 and next_batch_start < len(person_ids):
            sleep(config.pause_s)

    return results

def attempt_period(
    country: str,
    base_query: str,
    start_year: int,
    end_year: int,
    config: PEPScraperConfig | None = None,
    collected_ids: list[str] | None = None,
    tracker: QueryProgressTracker | None = None,
) -> list[dict[str, str]] | None:
    """Attempt to run a SPARQL query for a specific birth year window.

    Args:
        country (str): The country code or QID to filter results by nationality.
        base_query (str): The SPARQL query template to execute.
        start_year (int): The starting year of the birth date filter.
        end_year (int): The ending year of the birth date filter.
        config (PEPScraperConfig | None): Optional scraper configuration.
        collected_ids (list[str] | None): Optional list of QIDs for per-person queries.
        tracker (QueryProgressTracker | None): Optional progress tracker for attempts.

    Returns:
        list[dict[str, str]] | None: Result bindings when successful, otherwise ``None``.

    Raises:
        None.

    """
    active_config = config or PEPScraperConfig()

    if base_query == MAIN_QUERY:
        main_results = _execute_main_query(
            country,
            base_query,
            start_year,
            end_year,
            active_config,
            tracker,
        )
        return main_results if main_results else None

    if not collected_ids:
        return None

    person_results = _execute_person_queries(
        base_query,
        active_config,
        collected_ids,
        tracker,
    )
    return person_results if person_results else None


def main_query_format(
        start_year: int,
        end_year: int,
        base_query: str,
        country:str,
        )-> str:
    """Format SPARQL query, changing the details for eachh search parameter.

    Args:
        country (str): The country code or QID to filter results by nationality.
        start_year (int): The starting year of the birth date filter.
        end_year (int): The ending year of the birth date filter.
        config (PEPScraperConfig | None): Optional scraper configuration.
        base_query (str): Query being used for each loop
        
    Returns:
        str: A formatted SPARQL query that can be put to Wikidata .

    """
    country_id = get_country_id(country)
    if start_year == 0 and end_year == 0:
        if "date_filter" in base_query:
            query = base_query.replace(
            "?person wdt:P569 ?birthDate . date_filter",
            "FILTER NOT EXISTS { ?person wdt:P569 ?birthDate . }",
        ).replace("nationality_qid", country_id)
        else:
            query = base_query.replace("?person wdt:P569 ?birthDate .",
                                       "FILTER NOT EXISTS { ?person wdt:P569 ?birthDate . }")
    else:
        logger.info("Attempting query for period: %d to %d", start_year, end_year)
        if "date_filter" in base_query:
            query = base_query.replace(
                "date_filter",
                f'FILTER ("{start_year}-01-01T00:00:00Z"^^xsd:dateTime <= ?birthDate && '
                f'?birthDate < "{end_year}-01-01T00:00:00Z"^^xsd:dateTime).',
            ).replace("nationality_qid", country_id)
        else:
            query = base_query.replace("nationality_qid", country_id)
    

    return query

def bring_id_to_front(df: pd.DataFrame) -> pd.DataFrame:
    """Ensure the ID column is positioned first.

    Args:
        df (pd.DataFrame): DataFrame to reorder.

    Returns:
        pd.DataFrame: Reordered DataFrame with ID first when present.

    """
    if "ID" not in df.columns:
        return df
    ordered_columns = ["ID"] + [col for col in df.columns if col != "ID"]
    return df[ordered_columns]

def scrape_living_politicians(
    country: str,
    output_file: str | Path | None = None,
    config: PEPScraperConfig | None = None,
) -> tuple[int, str]:
    """Scrape only living politicians from a country.
    
    Args:
        country (str): Country code (e.g., 'uk', 'us') or Wikidata ID.
        output_file (str | Path | None, optional): Output CSV file path.
        config (PEPScraperConfig | None, optional): Configuration.
        
    Returns:
        tuple[int, str]: Number of records scraped and output file path.

    """
    logger.info("Using decade-based querying for living politicians")
    return scrape_country_politicians_by_decade(
        country=country,
        output_file=output_file,
        living_only=True,
        apply_relevance_filter=False,  # Already filtered by living status
        config=config,
    )
